"""
Gestion des commandes Excel
────────────────────────────
Deux fichiers Excel distincts selon la complexité :
  - orders/commandes_simples.xlsx   → commandes ≤ seuil de complexité
  - orders/commandes_complexes.xlsx → commandes > seuil (traitement manuel requis)

Chaque ligne représente une commande complète avec les infos client,
les articles, le total et le statut de paiement.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..config import settings

logger = logging.getLogger(__name__)

# ── Colonnes ──────────────────────────────────────────────────────────────────
HEADERS = ["Réf.", "Date", "Heure", "Nom", "Prénom", "Téléphone",
           "Articles", "Total (€)", "Paiement", "Statut"]
COLUMN_WIDTHS = [10, 12, 8, 16, 16, 16, 50, 12, 20, 22]

COLOR_SIMPLE  = "2E7D32"   # Vert foncé
COLOR_COMPLEX = "B71C1C"   # Rouge foncé
COLOR_WHITE   = "FFFFFF"


def _create_header(ws, color: str):
    header_font = Font(bold=True, color=COLOR_WHITE, size=11)
    header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    thin = Border(
        bottom=Side(style="thin", color=COLOR_WHITE),
        right=Side(style="thin", color=COLOR_WHITE),
    )
    for col, (header, width) in enumerate(zip(HEADERS, COLUMN_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def _get_or_create_workbook(filepath: str, is_complex: bool) -> Workbook:
    path = Path(filepath)
    if path.exists():
        return load_workbook(filepath)
    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes"
    _create_header(ws, COLOR_COMPLEX if is_complex else COLOR_SIMPLE)
    return wb


def write_order(
    order_items: List[dict],
    is_complex: bool = False,
    customer_lastname: str = "",
    customer_firstname: str = "",
    customer_phone: str = "",
    payment_method: str = "Non renseigné",
    payment_status: str = "En attente",
    total: float = 0.0,
    order_id: str = "",
):
    """
    Enregistre une commande dans le fichier Excel approprié.
    Une ligne par commande (les articles sont agrégés en une cellule).

    Args:
        order_items       : liste de {"produit": str, "quantite": int}
        is_complex        : True → commandes_complexes.xlsx
        customer_lastname : nom de famille du client
        customer_firstname: prénom du client
        customer_phone    : numéro de téléphone
        payment_method    : "CB" | "liquide" | "Non renseigné"
        payment_status    : "Payé par CB" | "Règlement sur place" | "Paiement refusé → sur place"
        total             : montant total en €
        order_id          : référence unique (ex : "A1B2C3D4")
    """
    orders_dir = Path(settings.orders_dir)
    orders_dir.mkdir(parents=True, exist_ok=True)

    filename = "commandes_complexes.xlsx" if is_complex else "commandes_simples.xlsx"
    filepath = str(orders_dir / filename)

    wb = _get_or_create_workbook(filepath, is_complex)
    ws = wb.active

    now = datetime.now()
    date_str = now.strftime("%d/%m/%Y")
    time_str = now.strftime("%H:%M")

    # Résumé des articles en une seule cellule
    articles_str = " | ".join(
        f"{item.get('quantite', 1)}× {item.get('produit', '?')}"
        for item in order_items
    )
    statut = "À traiter manuellement" if is_complex else "Confirmée"

    row = ws.max_row + 1
    bg = "F5F5F5" if row % 2 == 0 else "FFFFFF"
    row_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

    values = [
        order_id,
        date_str,
        time_str,
        customer_lastname,
        customer_firstname,
        customer_phone,
        articles_str,
        f"{total:.2f}" if total else "",
        payment_status,
        statut,
    ]

    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = row_fill
        # Articles (col 7) et paiement (col 9) alignés à gauche
        cell.alignment = Alignment(
            horizontal="left" if col in (7, 9) else "center",
            wrap_text=(col == 7),
        )

    wb.save(filepath)
    logger.info(
        f"Commande {order_id} enregistrée dans {filepath} "
        f"({customer_firstname} {customer_lastname}, {total:.2f}€, {payment_status})"
    )
