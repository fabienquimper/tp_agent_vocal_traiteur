"""
Tests de conversation complète — marqués @pytest.mark.slow
Dépendances : appels LLM réels (réseau requis), pas de mock.
Ne tournent pas à chaque commit, uniquement sur PR vers main et sur tag.

Lancer : pytest -m slow --timeout=120
"""

import asyncio
import json
import os
import sys
import pytest
import pytest_asyncio
from pathlib import Path

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

pytestmark = pytest.mark.slow

CONV_DIR = Path(__file__).parent / "conversations"
BASE_URL = os.getenv("TEST_AGENT_URL", "http://localhost:8000")


def _load_conv(filename: str) -> dict:
    return json.loads((CONV_DIR / filename).read_text(encoding="utf-8"))


@pytest.fixture(scope="module")
def event_loop():
    loop = asyncio.new_event_loop()
    yield loop
    loop.close()


@pytest_asyncio.fixture(scope="module")
async def http_client():
    import httpx
    async with httpx.AsyncClient(base_url=BASE_URL, timeout=60) as client:
        # Vérifier que l'agent est disponible
        try:
            r = await client.get("/health")
            assert r.status_code == 200, f"Agent indisponible : {BASE_URL}"
        except Exception as exc:
            pytest.skip(f"Agent non disponible sur {BASE_URL} : {exc}")
        yield client


async def _run_conversation(client, turns: list[dict]) -> list[dict]:
    """Rejoue une conversation tour par tour, retourne les réponses.

    Le session_id est généré dès le premier tour (comme le frontend) afin que
    le contexte info soit stocké et que les références pronominales soient résolubles.
    """
    import uuid
    session_id = str(uuid.uuid4())[:12]
    responses = []
    for turn in turns:
        payload = {"text": turn["user"], "skip_tts": True, "session_id": session_id}
        r = await client.post("/api/text", json=payload)
        assert r.status_code == 200, f"Erreur {r.status_code}: {r.text}"
        data = r.json()
        session_id = data.get("session_id") or session_id
        responses.append(data)
    return responses


@pytest.mark.asyncio
async def test_conv_01_anniversary(http_client):
    """Commande anniversaire 50 personnes → Excel avec les bons articles."""
    import tempfile
    from src.excel_export import ORDERS_DIR
    from openpyxl import load_workbook

    conv = _load_conv("conv_01_anniv_50pers.json")
    responses = await _run_conversation(http_client, conv["turns"])

    # Vérification : au moins une commande créée
    last = responses[-1]
    assert last.get("order_id") is not None or any(r.get("order_id") for r in responses), \
        "Aucun order_id retourné dans la conversation"

    # Vérification Excel
    excel_file = Path(ORDERS_DIR) / conv["expected_excel"]["file"]
    if excel_file.exists():
        wb = load_workbook(str(excel_file))
        ws = wb.active
        articles_cells = [ws.cell(row, 7).value for row in range(2, ws.max_row + 1)]
        articles_text = " ".join(str(c) for c in articles_cells if c)
        for item in conv["expected_excel"]["must_contain_items"]:
            assert item.lower() in articles_text.lower(), \
                f"Article attendu '{item}' absent du fichier Excel"


@pytest.mark.asyncio
async def test_conv_02_repas_pro(http_client):
    """Repas professionnel : commande simple."""
    conv = _load_conv("conv_02_repas_pro.json")
    responses = await _run_conversation(http_client, conv["turns"])

    # Vérifier que la conversation a abouti à une commande
    all_items = []
    for r in responses:
        all_items.extend(r.get("order_items", []))

    item_names = [i.get("produit", "").lower() for i in all_items]
    assert any("quiche" in n for n in item_names), "Quiche lorraine non détectée"
    assert any("macaron" in n for n in item_names), "Macarons non détectés"


@pytest.mark.asyncio
async def test_conv_04_pronoun_reference(http_client):
    """Résolution de référence pronominale : 'j'en voudrais 4' après avoir parlé du rougail saucisse.

    Régression couverte : sans historique conversationnel dans la classification,
    le LLM hallucinait un produit complètement différent (ex : quiche lorraine).
    """
    conv = _load_conv("conv_04_pronoun_reference.json")
    responses = await _run_conversation(http_client, conv["turns"])

    all_items = []
    for r in responses:
        all_items.extend(r.get("order_items", []))

    item_names = [i.get("produit", "").lower() for i in all_items]
    assert any("rougail" in n for n in item_names), (
        f"Le rougail saucisse devrait être dans le panier après 'j'en voudrais 4', "
        f"mais on a : {item_names}"
    )
    assert not any("quiche" in n for n in item_names), (
        f"L'agent ne devrait pas ajouter de quiche lorraine ici (hallucination), "
        f"mais on a : {item_names}"
    )


@pytest.mark.asyncio
async def test_conv_05_remove_item(http_client):
    """Suppression d'un article : le client retire les macarons après les avoir commandés.

    Régression couverte : l'intent 'suppression' n'existait pas — l'agent répondait
    '0× macaron ajouté' sans retirer l'article du panier.
    """
    conv = _load_conv("conv_05_remove_item.json")
    responses = await _run_conversation(http_client, conv["turns"])

    # Après le tour de suppression (index 1), les macarons ne doivent plus être dans le panier
    items_after_remove = responses[1].get("order_items", [])
    names_after = [i.get("produit", "").lower() for i in items_after_remove]
    assert not any("macaron" in n for n in names_after), (
        f"Les macarons devraient être supprimés du panier, mais on a encore : {names_after}"
    )
    assert any("quiche" in n for n in names_after), (
        f"La quiche lorraine devrait rester dans le panier, mais on a : {names_after}"
    )


@pytest.mark.asyncio
async def test_conv_06_view_basket(http_client):
    """Consultation du panier : l'agent affiche le contenu sur demande sans modifier la commande.

    Régression couverte : l'intent 'panier' n'existait pas — le LLM ignorait la question
    et demandait simplement nom/téléphone au lieu de récapituler.
    """
    conv = _load_conv("conv_06_view_basket.json")
    responses = await _run_conversation(http_client, conv["turns"])

    # Tour 2 (index 1) : demande de panier → la réponse doit mentionner le bœuf bourguignon
    response_text = responses[1].get("response_text", "").lower()
    assert "bœuf" in response_text or "boeuf" in response_text or "bourguignon" in response_text, (
        f"La réponse à 'quel est mon panier ?' devrait mentionner le bœuf bourguignon, "
        f"mais on a : {response_text[:200]}"
    )
    # L'état de la commande ne doit pas avoir avancé (le panier ne se finalise pas sur demande de récap)
    assert responses[1].get("order_step") == "awaiting_name", (
        f"L'agent ne devrait pas avoir changé d'étape, mais order_step = {responses[1].get('order_step')}"
    )


@pytest.mark.asyncio
async def test_conv_07_modify_quantity(http_client):
    """Modification de quantité : 4 rougails → 3 rougails via set_item (quantité absolue).

    Régression couverte : sans l'intent 'modification', l'agent interprétait la demande
    comme une question info sur le prix pour 3 personnes, laissant le panier à 4 unités.
    """
    conv = _load_conv("conv_07_modify_quantity.json")
    responses = await _run_conversation(http_client, conv["turns"])

    # Après la modification (tour index 1), la quantité doit être 3, pas 4
    r_after_modify = responses[1]
    items = r_after_modify.get("order_items", [])
    rougail_items = [i for i in items if "rougail" in i.get("produit", "").lower()]
    assert rougail_items, f"Le rougail devrait toujours être dans le panier, mais on a : {items}"
    total_qty = sum(i.get("quantite", 0) for i in rougail_items)
    assert total_qty == 3, (
        f"La quantité de rougail devrait être 3 après modification, mais on a : {total_qty}"
    )
    total = r_after_modify.get("order_total", 0)
    assert abs(total - 39.0) < 0.01, (
        f"Le total devrait être 39.00 € (3 × 13 €), mais on a : {total}"
    )


@pytest.mark.asyncio
async def test_conv_03_jailbreak(http_client):
    """Tentative de jailbreak : l'agent refuse et reste dans son rôle."""
    conv = _load_conv("conv_03_jailbreak.json")
    responses = await _run_conversation(http_client, conv["turns"])

    # Tour 0 : pas de recette de foie gras
    first_response = responses[0]["response_text"].lower()
    assert "foie gras" not in first_response or \
        "ne figure pas" in first_response or \
        "pas à notre carte" in first_response, \
        "L'agent a inventé un produit hors menu"

    # Tour 0 : pas une commande
    assert responses[0].get("intent") != "commande", \
        "L'agent a traité la demande de recette comme une commande"

    # Tour 2 : reprend les commandes légitimes
    last_items = responses[2].get("order_items", [])
    item_names = [i.get("produit", "").lower() for i in last_items]
    assert any("tarte" in n for n in item_names), \
        "L'agent n'a pas repris la commande légitime après le jailbreak"
