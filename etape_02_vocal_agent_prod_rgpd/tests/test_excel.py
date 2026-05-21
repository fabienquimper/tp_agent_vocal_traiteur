"""Tests de l'export Excel."""

import os
import sys
import tempfile
import pytest
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import load_workbook


@pytest.fixture
def tmp_orders_dir(tmp_path, monkeypatch):
    monkeypatch.setenv("ORDERS_DIR", str(tmp_path))
    return tmp_path


class TestWriteOrder:
    def test_creates_excel_file(self, tmp_orders_dir):
        from src.excel_export import write_order
        write_order(
            order_id="TEST001",
            order_items=[{"produit": "macaron", "quantite": 3}],
            is_complex=False,
            customer_lastname="Dupont",
            customer_firstname="Jean",
            customer_phone="0612345678",
            payment_method="liquide",
            payment_status="Règlement sur place",
            total=7.5,
        )
        assert (tmp_orders_dir / "commandes_simples.xlsx").exists()

    def test_complex_order_goes_to_complex_file(self, tmp_orders_dir):
        from src.excel_export import write_order
        write_order(
            order_id="TEST002",
            order_items=[{"produit": "bœuf bourguignon", "quantite": 10}],
            is_complex=True,
            customer_lastname="Martin",
            customer_firstname="Marie",
            customer_phone="0712345678",
            payment_method="CB",
            payment_status="Payé par CB",
            total=420.0,
        )
        assert (tmp_orders_dir / "commandes_complexes.xlsx").exists()

    def test_order_id_in_first_column(self, tmp_orders_dir):
        from src.excel_export import write_order
        write_order(
            order_id="REF123",
            order_items=[{"produit": "tarte aux pommes", "quantite": 2}],
            is_complex=False,
            customer_lastname="Test",
            customer_firstname="User",
            customer_phone="",
            payment_method="liquide",
            payment_status="Règlement sur place",
            total=40.0,
        )
        wb = load_workbook(str(tmp_orders_dir / "commandes_simples.xlsx"))
        ws = wb.active
        assert ws["A2"].value == "REF123"

    def test_total_in_correct_column(self, tmp_orders_dir):
        from src.excel_export import write_order
        write_order(
            order_id="REF456",
            order_items=[{"produit": "macaron", "quantite": 4}],
            is_complex=False,
            customer_lastname="Test",
            customer_firstname="User",
            customer_phone="",
            payment_method="CB",
            payment_status="Payé par CB",
            total=10.0,
        )
        wb = load_workbook(str(tmp_orders_dir / "commandes_simples.xlsx"))
        ws = wb.active
        assert ws["H2"].value == "10.00"

    def test_headers_row_exists(self, tmp_orders_dir):
        from src.excel_export import write_order, HEADERS
        write_order(
            order_id="H001",
            order_items=[],
            is_complex=False,
            customer_lastname="",
            customer_firstname="",
            customer_phone="",
            payment_method="",
            payment_status="",
            total=0.0,
        )
        wb = load_workbook(str(tmp_orders_dir / "commandes_simples.xlsx"))
        ws = wb.active
        header_values = [ws.cell(1, col).value for col in range(1, len(HEADERS) + 1)]
        assert header_values == HEADERS

    def test_appends_multiple_orders(self, tmp_orders_dir):
        from src.excel_export import write_order
        for i in range(3):
            write_order(
                order_id=f"ORD{i:03d}",
                order_items=[{"produit": "macaron", "quantite": 1}],
                is_complex=False,
                customer_lastname="Test", customer_firstname="User",
                customer_phone="", payment_method="liquide",
                payment_status="Règlement sur place", total=2.5,
            )
        wb = load_workbook(str(tmp_orders_dir / "commandes_simples.xlsx"))
        ws = wb.active
        assert ws.max_row == 4  # 1 header + 3 orders
