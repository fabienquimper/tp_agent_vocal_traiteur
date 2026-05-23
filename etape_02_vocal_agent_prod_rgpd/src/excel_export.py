"""
Export Excel des commandes — Traiteur Dupont Étape 02

Deux fichiers séparés selon la complexité :
  - orders/commandes_simples.xlsx   → commandes ≤ seuil de complexité
  - orders/commandes_complexes.xlsx → commandes > seuil (traitement manuel requis)

Les données personnelles (nom, téléphone) sont écrites uniquement dans ces fichiers locaux,
jamais dans les logs ni transmises à un service externe.
"""

import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

def _get_orders_dir() -> str:
    return os.getenv("ORDERS_DIR", "/app/orders")

ORDERS_DIR = _get_orders_dir()

HEADERS = ["Réf.", "Date", "Heure", "Nom", "Prénom", "Téléphone", "Articles", "Total (€)", "Paiement", "Statut"]
COLUMN_WIDTHS = [10, 12, 8, 16, 16, 16, 50, 12, 22, 22]

_COLOR_SIMPLE = "2E7D32"
_COLOR_COMPLEX = "B71C1C"
_COLOR_WHITE = "FFFFFF"


def _ensure_workbook(filepath: Path, is_complex: bool) -> Workbook:
    if filepath.exists():
        return load_workbook(str(filepath))
    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes"
    _write_header(ws, _COLOR_COMPLEX if is_complex else _COLOR_SIMPLE)
    return wb


def _write_header(ws, color: str) -> None:
    font = Font(bold=True, color=_COLOR_WHITE, size=11)
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    border = Border(
        bottom=Side(style="thin", color=_COLOR_WHITE),
        right=Side(style="thin", color=_COLOR_WHITE),
    )
    for col, (header, width) in enumerate(zip(HEADERS, COLUMN_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def write_order(
    order_id: str,
    order_items: list[dict],
    is_complex: bool,
    customer_lastname: str,
    customer_firstname: str,
    customer_phone: str,
    payment_method: str,
    payment_status: str,
    total: float,
) -> None:
    """Enregistre une commande dans le fichier Excel approprié."""
    orders_path = Path(_get_orders_dir())
    orders_path.mkdir(parents=True, exist_ok=True)

    filename = "commandes_complexes.xlsx" if is_complex else "commandes_simples.xlsx"
    filepath = orders_path / filename

    wb = _ensure_workbook(filepath, is_complex)
    ws = wb.active

    now = datetime.now()
    articles_str = " | ".join(
        f"{item.get('quantite', 1)}× {item.get('produit', '?')}" for item in order_items
    )
    statut = "À traiter manuellement" if is_complex else "Confirmée"

    row = ws.max_row + 1
    bg = "F5F5F5" if row % 2 == 0 else "FFFFFF"
    row_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")

    values = [
        order_id,
        now.strftime("%d/%m/%Y"),
        now.strftime("%H:%M"),
        customer_lastname,
        customer_firstname,
        customer_phone,
        articles_str,
        f"{total:.2f}" if total else "",
        payment_status,
        statut,
    ]
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = row_fill
        cell.alignment = Alignment(
            horizontal="left" if col in (7, 9) else "center",
            wrap_text=(col == 7),
        )

    wb.save(str(filepath))
