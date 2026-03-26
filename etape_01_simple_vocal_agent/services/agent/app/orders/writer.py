"""
Gestion des commandes Excel
────────────────────────────
Deux fichiers Excel distincts selon la complexité :
  - orders/commandes_simples.xlsx   → commandes ≤ seuil de complexité
  - orders/commandes_complexes.xlsx → commandes > seuil (traitement manuel requis)

Chaque fichier est créé automatiquement avec un en-tête stylisé
s'il n'existe pas encore, puis les nouvelles lignes sont ajoutées.
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..config import settings

logger = logging.getLogger(__name__)

# ── Constantes ────────────────────────────────────────────────────────────────
HEADERS = ["Date", "Heure", "Produit", "Quantité", "Statut", "Source"]
COLUMN_WIDTHS = [14, 10, 35, 12, 18, 15]

COLOR_SIMPLE = "2E7D32"    # Vert foncé – en-tête commandes simples
COLOR_COMPLEX = "B71C1C"   # Rouge foncé – en-tête commandes complexes
COLOR_WHITE = "FFFFFF"


def _create_header(ws, color: str):
    """Crée la ligne d'en-tête avec style."""
    header_font = Font(bold=True, color=COLOR_WHITE, size=11)
    header_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color=COLOR_WHITE),
        right=Side(style="thin", color=COLOR_WHITE),
    )

    for col, (header, width) in enumerate(zip(HEADERS, COLUMN_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"  # Fige la ligne d'en-tête


def _get_or_create_workbook(filepath: str, is_complex: bool) -> Workbook:
    """Charge le classeur existant ou en crée un nouveau."""
    path = Path(filepath)
    if path.exists():
        return load_workbook(filepath)

    wb = Workbook()
    ws = wb.active
    ws.title = "Commandes"
    color = COLOR_COMPLEX if is_complex else COLOR_SIMPLE
    _create_header(ws, color)
    return wb


def write_order(order_items: List[dict], is_complex: bool = False):
    """
    Enregistre les articles de la commande dans le fichier Excel approprié.

    Args:
        order_items : liste de {"produit": str, "quantite": int}
        is_complex  : True → commandes_complexes.xlsx, False → commandes_simples.xlsx
    """
    # Création du dossier orders/ si absent
    orders_dir = Path(settings.orders_dir)
    orders_dir.mkdir(parents=True, exist_ok=True)

    filename = "commandes_complexes.xlsx" if is_complex else "commandes_simples.xlsx"
    filepath = str(orders_dir / filename)

    wb = _get_or_create_workbook(filepath, is_complex)
    ws = wb.active

    now = datetime.now()
    date_str = now.strftime("%d/%m/%Y")
    time_str = now.strftime("%H:%M")
    statut = "À traiter manuellement" if is_complex else "En attente confirmation"
    source = "Assistant vocal"

    # Alternance de couleurs pour les lignes (lisibilité)
    next_row = ws.max_row + 1
    for i, item in enumerate(order_items):
        row = next_row + i
        bg_color = "F5F5F5" if row % 2 == 0 else "FFFFFF"
        row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

        values = [
            date_str,
            time_str,
            item.get("produit", ""),
            item.get("quantite", 0),
            statut,
            source,
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="center" if col != 3 else "left")

    wb.save(filepath)
    logger.info(
        f"{'Commande complexe' if is_complex else 'Commande simple'} "
        f"enregistrée dans {filepath} : {order_items}"
    )
